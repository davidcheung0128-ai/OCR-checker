# backend/excel_exporter.py
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO

def generate_pressure_drop_excel(
    company_name: str = "Young's Engineering Company Limited",
    project_name: str = "Dedicated Rehousing at Ma Tau Kok",
    location: str = "B1/F Master Water Meter Room",
    ref_no: str = "EAF-B1-02",
    flow_rate: float = 0.25,
    specified_esp: float = 400.0,
    offered_esp: float = 450.0,
    sections_data: list = None
) -> BytesIO:
    """
    接收計算完成的 sections_data 陣列，動態生成 Young's Engineering 企業標準 Excel 計算書。
    回傳 BytesIO 流供 API 直接供前端下載。
    """
    if sections_data is None:
        sections_data = []

    wb = Workbook()
    ws = wb.active
    ws.title = "ESP Calculation"
    ws.views.sheetView[0].showGridLines = True

    # 樣式定義
    title_font = Font(name="Calibri", size=14, bold=True)
    header_font = Font(name="Calibri", size=10, bold=True)
    data_font = Font(name="Calibri", size=10)
    bold_font = Font(name="Calibri", size=10, bold=True)

    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='A6A6A6'),
        right=Side(style='thin', color='A6A6A6'),
        top=Side(style='thin', color='A6A6A6'),
        bottom=Side(style='thin', color='A6A6A6')
    )

    # 1. 公司與專案表頭資訊
    ws['A1'] = company_name
    ws['A1'].font = title_font
    ws['A2'] = project_name
    ws['A2'].font = Font(name="Calibri", size=12, italic=True)

    ws['A4'] = "External Static Pressure Calculation"
    ws['A4'].font = title_font

    ws['A6'] = "Location :"
    ws['C6'] = location
    ws['N6'] = "Specified"
    ws['O6'] = "Offered"
    ws['N6'].font = header_font
    ws['O6'].font = header_font

    ws['A7'] = "Ref. No :"
    ws['C7'] = ref_no
    ws['M7'] = "Flow Rate (m3/s) :"
    ws['N7'] = flow_rate
    ws['O7'] = flow_rate

    ws['A8'] = "Area Served :"
    ws['C8'] = location
    ws['M8'] = "External Static Pressure (Pa) :"
    ws['N8'] = specified_esp
    ws['O8'] = offered_esp

    # 2. 表格標頭 (Row 10 - 12)
    headers_row10 = {
        "A10": "Section", "B10": "Air Qty.", "C10": "Duct Section", "F10": "Type of Fitting",
        "G10": "Main Duct Size", "K10": "Hydraulic Diameter", "L10": "Reynolds Number",
        "M10": "Roughness Factor", "N10": "Darcy Friction", "O10": "Total Length",
        "P10": "Friction Factor", "Q10": "Fitting Loss Coeff.", "R10": "Velocity (m/s)",
        "S10": "Velocity Pressure (Pa)", "T10": "Friction Loss (Pa)"
    }

    for cell_ref, text in headers_row10.items():
        ws[cell_ref] = text
        ws[cell_ref].font = header_font
        ws[cell_ref].fill = header_fill
        ws[cell_ref].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 3. 填入數據列
    current_row = 13
    suction_header_written = False
    discharge_header_written = False

    for idx, sec in enumerate(sections_data, start=1):
        sec_type = sec.get("type", "Suction").capitalize()

        if sec_type == "Suction" and not suction_header_written:
            ws.cell(row=current_row, column=1, value="Suction").font = bold_font
            current_row += 1
            suction_header_written = True
        elif sec_type == "Discharge" and not discharge_header_written:
            ws.cell(row=current_row, column=1, value="Discharge").font = bold_font
            current_row += 1
            discharge_header_written = True

        ws.cell(row=current_row, column=1, value=idx)                             # Section ID
        ws.cell(row=current_row, column=2, value=sec.get("flow_rate", flow_rate))  # Air Qty
        ws.cell(row=current_row, column=6, value=sec.get("fitting_name", "Run"))   # Type of Fitting
        ws.cell(row=current_row, column=7, value=sec.get("a_mm", 0))               # a
        ws.cell(row=current_row, column=8, value="x")
        ws.cell(row=current_row, column=9, value=sec.get("b_mm", 0))               # b
        ws.cell(row=current_row, column=11, value=sec.get("D_mm", 0))              # Hydraulic Diameter
        ws.cell(row=current_row, column=12, value=sec.get("Re", 0))                # Reynolds Number
        ws.cell(row=current_row, column=13, value=0.15)                            # Roughness
        ws.cell(row=current_row, column=14, value=sec.get("f", 0.02))              # Darcy f
        ws.cell(row=current_row, column=15, value=sec.get("length_m", 0.0))        # Length L
        ws.cell(row=current_row, column=17, value=sec.get("c_coefficient", 0.0))   # Loss Coeff
        ws.cell(row=current_row, column=18, value=sec.get("velocity_ms", 0.0))     # Velocity
        ws.cell(row=current_row, column=19, value=sec.get("velocity_pressure_pa", 0.0)) # Vel Pressure
        ws.cell(row=current_row, column=20, value=sec.get("total_pressure_loss_pa", 0.0)) # Pressure Loss

        for c in range(1, 21):
            cell = ws.cell(row=current_row, column=c)
            cell.font = data_font
            cell.border = thin_border
            if c not in [6, 8]:
                cell.alignment = Alignment(horizontal="right")

        current_row += 1

    # 4. 總壓降與安全係數計算
    current_row += 1
    ws.cell(row=current_row, column=17, value="Total Pressure Loss").font = bold_font
    ws.cell(row=current_row, column=20, value=f"=SUM(T13:T{current_row-2})").font = bold_font

    current_row += 1
    ws.cell(row=current_row, column=17, value="Safety Factor : 20%").font = bold_font
    ws.cell(row=current_row, column=20, value=f"=T{current_row-1}*0.20").font = bold_font

    current_row += 1
    ws.cell(row=current_row, column=17, value="Grand Total (Pa)").font = bold_font
    ws.cell(row=current_row, column=20, value=f"=T{current_row-2}+T{current_row-1}").font = bold_font

    # 自動調整欄寬
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output